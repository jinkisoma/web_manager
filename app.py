import os
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, Response, jsonify
from pymongo import MongoClient
import certifi
from bson.objectid import ObjectId
import pandas as pd
import io
from datetime import datetime, timedelta
from urllib.parse import quote
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)
app.secret_key = 'your_very_secret_key' # 실제 운영시에는 더 복잡한 키를 사용하세요.

# --- MongoDB 연결 설정 ---
# Atlas에서 복사한 연결 문자열을 여기에 붙여넣으세요. <username>, <password>를 실제 정보로 수정해야 합니다.
MONGO_URI = os.environ.get('DATABASE_URL', "mongodb+srv://jinkikim0629_db_user:i5n9VjN6sngz2pc6@cluster0.jzo4itl.mongodb.net/?appName=Cluster0")
client = MongoClient(MONGO_URI, tlsCAFile=certifi.where(), tlsDisableOCSPStapling=True)
db = client.settlement_note # 데이터베이스 이름 설정 (없으면 자동 생성)
users_collection = db.user_data # 컬렉션(테이블과 유사) 이름 설정

ATTACHMENT_DIR = "attachments"

# --- 거래처 및 작업 데이터 관리 (이 부분은 DB에 저장하거나 그대로 유지할 수 있습니다) ---
CLIENT_WORK_DATA = { # 이 데이터는 이제 API를 통해 클라이언트에 제공됩니다.
    '로지비': {
        '라벨작업': {'content': '단상자 바코드작업', 'price': 100},
        '포장작업': {'content': '선물세트 포장', 'price': 500}
    },

    '비플레인': {
        '소분작업': {'content': '샘플 소분', 'price': 50},
        '검수작업': {'content': '제품 외관 검수', 'price': 80}
    },


    '릴라이블': {
 '소분작업': {'content': '샘플 소분', 'price': 50},
  '검수작업': {'content': '제품 외관 검수', 'price': 80},
    '라벨작업': {'content': '단상자 바코드작업', 'price': 100},
        '포장작업': {'content': '선물세트 포장', 'price': 500}
    },

    '릴라이블(대성)': {
 '소분작업': {'content': '샘플 소분', 'price': 50},
  '검수작업': {'content': '제품 외관 검수', 'price': 80},
    '라벨작업': {'content': '단상자 바코드작업', 'price': 100},
        '포장작업': {'content': '선물세트 포장', 'price': 500}
    },

    '릴라이블(랩)': {
 '소분작업': {'content': '샘플 소분', 'price': 50},
  '검수작업': {'content': '제품 외관 검수', 'price': 80},
    '라벨작업': {'content': '단상자 바코드작업', 'price': 100},
        '포장작업': {'content': '아웃박스를 열여서 유관검수후에 다시 재포장해서 닫는작업', 'price': 1000}

    }
}

# 엑셀 헤더 한글 매핑
HEADER_MAP = {
    'id': '고유번호',
    'work_date': '작업일자',
    'client': '거래처',
    'author': '작성자',
    'product_code': '업체상품코드',
    'work_type': '작업 구분',
    'content': '내용',
    'product_name': '상품명',
    'quantity': '작업수량',
    'box_quantity': '박스수량',
    'unit_price': '금액(단가)',
    'total_amount': '합계',
    'remarks': '주문자',
    'tracking_number': '송장번호',
    'confirmed': '확정여부'
}

# 확정 취소 및 관리자 비밀번호
CONFIRM_CANCEL_PASSWORD = "1234"
ADMIN_OVERRIDE_PASSWORD = "2580"

# 첨부파일 디렉토리 생성
def init_dirs():
    if not os.path.exists(ATTACHMENT_DIR):
        os.makedirs(ATTACHMENT_DIR)

@app.template_filter('number_format')
def number_format(value):
    if value is None:
        return ""
    try:
        return format(int(value), ',')
    except (ValueError, TypeError):
        return value

@app.route('/')
def index():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    author_filter = request.args.get('author_filter')
    keyword = request.args.get('keyword')

    query = {}
    
    if start_date_str and end_date_str:
        query['work_date'] = {'$gte': start_date_str, '$lte': end_date_str}

    if author_filter:
        query['author'] = author_filter

    if keyword:
        query['$or'] = [
            {'client': {'$regex': keyword, '$options': 'i'}},
            {'author': {'$regex': keyword, '$options': 'i'}},
            {'product_name': {'$regex': keyword, '$options': 'i'}},
            {'content': {'$regex': keyword, '$options': 'i'}},
            {'tracking_number': {'$regex': keyword, '$options': 'i'}},
        ]

    # MongoDB에서 데이터 조회 (최신순으로 정렬)
    users = list(users_collection.find(query).sort([('work_date', -1), ('_id', -1)]))

    # ObjectId를 문자열로 변환 (JSON 직렬화 및 URL 생성을 위해)
    for user in users:
        user['id'] = str(user['_id'])

    total_count = len(users)
    confirmed_count = sum(1 for user in users if user.get('confirmed'))
    unconfirmed_count = total_count - confirmed_count

    # 거래처 목록은 DB에서 고유값으로 가져오기
    clients = list(CLIENT_WORK_DATA.keys())
    db_clients = users_collection.distinct('client')
    clients = sorted(list(set(clients + db_clients)))

    return render_template('index.html', users=users,
                           today_date=datetime.now().strftime('%Y-%m-%d'),
                           start_date=start_date_str, end_date=end_date_str, keyword=keyword,
                           author_filter=author_filter, clients=clients,
                           total_count=total_count, confirmed_count=confirmed_count, unconfirmed_count=unconfirmed_count)

@app.route('/api/work-items/<client_name>')
def get_work_items(client_name):
    # 이 API는 기존과 동일하게 작동
    work_items = CLIENT_WORK_DATA.get(client_name, {})
    return jsonify(work_items)

@app.route('/add', methods=['POST'])
def add_user():
    try:
        client = request.form.get('client_select')
        if client == 'direct':
            client = request.form.get('client_direct')

        work_type = request.form.get('work_type_select')
        if work_type == 'direct':
            work_type = request.form.get('work_type_direct')
        
        orderer = request.form.get('orderer_select')
        if orderer == 'direct':
            orderer = request.form.get('orderer_direct')

        attachment_filename = None
        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                attachment_filename = file.filename
                file.save(os.path.join(ATTACHMENT_DIR, attachment_filename))

        quantity = int(request.form.get('quantity') or 0)
        unit_price = int(request.form.get('unit_price') or 0)
        total_amount = quantity * unit_price

        new_user = {
            "work_date": request.form.get('work_date'),
            "client": client,
            "author": request.form.get('author'),
            "product_code": request.form.get('product_code'),
            "tracking_number": request.form.get('tracking_number'),
            "work_type": work_type,
            "content": request.form.get('content'),
            "product_name": request.form.get('product_name'),
            "quantity": quantity,
            "box_quantity": int(request.form.get('box_quantity') or 0),
            "unit_price": unit_price,
            "total_amount": total_amount,
            "attachment": attachment_filename,
            "remarks": orderer,
            "confirmed": False,
            "created_at": datetime.utcnow()
        }
        users_collection.insert_one(new_user)
        flash('데이터가 성공적으로 추가되었습니다.', 'success')
    except Exception as e:
        flash(f'오류가 발생했습니다: {e}', 'error')
    return redirect(url_for('index'))

@app.route('/edit/<id>')
def edit_form(id):
    user = users_collection.find_one({'_id': ObjectId(id)})
    
    if user is None:
        flash('해당 데이터를 찾을 수 없습니다.', 'error')
        return redirect(url_for('index'))
    
    user['id'] = str(user['_id'])
    is_readonly = user['confirmed']

    current_author = request.args.get('current_author', '')
    override_password = request.args.get('override_password', '')

    is_owner = user['author'] == current_author
    is_admin = override_password == ADMIN_OVERRIDE_PASSWORD

    if not is_readonly and not is_owner and not is_admin:
        flash('타인의 데이터를 수정할 권한이 없습니다.', 'error')
        return redirect(url_for('index'))

    if is_readonly:
        flash('확정된 데이터는 수정할 수 없습니다.', 'error')
    
    clients = list(CLIENT_WORK_DATA.keys())
    return render_template('edit.html', user=user, clients=clients, is_readonly=is_readonly)

@app.route('/update/<id>', methods=['POST'])
def update_user(id):
    current_author = request.form.get('current_author', '')
    override_password = request.form.get('override_password', '')

    record = users_collection.find_one({'_id': ObjectId(id)}, {'author': 1, 'confirmed': 1})

    is_owner = record and record['author'] == current_author
    is_admin = override_password == ADMIN_OVERRIDE_PASSWORD

    if record and (record['confirmed'] or (not is_owner and not is_admin)):
        flash('확정되었거나 권한이 없는 데이터는 수정할 수 없습니다.', 'error')
        return redirect(url_for('index'))
    
    try:
        client = request.form.get('client_select')
        if client == 'direct': client = request.form.get('client_direct')
        work_type = request.form.get('work_type_select')
        if work_type == 'direct': work_type = request.form.get('work_type_direct')
        orderer = request.form.get('orderer_select')
        if orderer == 'direct': orderer = request.form.get('orderer_direct')

        quantity = int(request.form.get('quantity') or 0)
        unit_price = int(request.form.get('unit_price') or 0)
        total_amount = quantity * unit_price

        update_data = {
            "work_date": request.form.get('work_date'),
            "client": client,
            "author": request.form.get('author'),
            "product_code": request.form.get('product_code'),
            "tracking_number": request.form.get('tracking_number'),
            "work_type": work_type,
            "content": request.form.get('content'),
            "product_name": request.form.get('product_name'),
            "quantity": quantity,
            "box_quantity": int(request.form.get('box_quantity') or 0),
            "unit_price": unit_price,
            "total_amount": total_amount,
            "remarks": orderer
        }

        attachment_filename = request.form.get('existing_attachment', '')
        if request.form.get('delete_attachment'):
            if attachment_filename and os.path.exists(os.path.join(ATTACHMENT_DIR, attachment_filename)):
                os.remove(os.path.join(ATTACHMENT_DIR, attachment_filename))
            attachment_filename = None

        if 'attachment' in request.files:
            file = request.files['attachment']
            if file.filename != '':
                if attachment_filename and os.path.exists(os.path.join(ATTACHMENT_DIR, attachment_filename)):
                    os.remove(os.path.join(ATTACHMENT_DIR, attachment_filename))
                attachment_filename = file.filename
                file.save(os.path.join(ATTACHMENT_DIR, attachment_filename))
        
        update_data['attachment'] = attachment_filename

        users_collection.update_one({'_id': ObjectId(id)}, {'$set': update_data})
        flash('데이터가 성공적으로 수정되었습니다.', 'success')
    except Exception as e:
        flash(f'수정 중 오류 발생: {e}', 'error')
    return redirect(url_for('index'))

@app.route('/delete/<id>')
def delete_user(id):
    current_author = request.args.get('current_author', '')
    override_password = request.args.get('override_password', '')
    record = users_collection.find_one({'_id': ObjectId(id)})

    is_owner = record and record['author'] == current_author
    is_admin = override_password == ADMIN_OVERRIDE_PASSWORD

    if record and (record['confirmed'] or (not is_owner and not is_admin)):
        flash('확정되었거나 권한이 없는 데이터는 삭제할 수 없습니다.', 'error')
        return redirect(request.referrer or url_for('index'))
    
    if record and record.get('attachment'):
        attachment_path = os.path.join(ATTACHMENT_DIR, record['attachment'])
        if os.path.exists(attachment_path):
            os.remove(attachment_path)

    users_collection.delete_one({'_id': ObjectId(id)})
    flash('데이터가 삭제되었습니다.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/confirm/<id>', methods=['POST'])
def confirm_user(id):
    current_author = request.form.get('current_author', '')
    override_password = request.form.get('override_password', '')
    record = users_collection.find_one({'_id': ObjectId(id)}, {'author': 1})

    is_owner = record and record['author'] == current_author
    is_admin = override_password == ADMIN_OVERRIDE_PASSWORD

    if not is_owner and not is_admin:
        flash('타인의 데이터를 확정할 권한이 없습니다.', 'error')
        return redirect(request.referrer or url_for('index'))

    users_collection.update_one({'_id': ObjectId(id)}, {'$set': {'confirmed': True}})
    flash(f'데이터가 확정 처리되었습니다.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/confirm_all', methods=['POST'])
def confirm_all():
    current_author = request.form.get('current_author')
    ids_to_confirm = request.form.getlist('confirm_ids')
    if not ids_to_confirm:
        flash('확정할 데이터가 없습니다.', 'error')
        return redirect(request.referrer or url_for('index'))

    if not current_author:
        flash('작성자 정보가 없어 전체 확정을 진행할 수 없습니다.', 'error')
        return redirect(request.referrer or url_for('index'))

    object_ids_to_confirm = [ObjectId(id) for id in ids_to_confirm]
    result = users_collection.update_many(
        {'_id': {'$in': object_ids_to_confirm}, 'author': current_author, 'confirmed': False},
        {'$set': {'confirmed': True}}
    )
    
    flash(f'{result.modified_count}개의 데이터가 전체 확정 처리되었습니다.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/unconfirm/<id>', methods=['POST'])
def unconfirm_user(id):
    password = request.form.get('password')
    current_author = request.form.get('current_author', '')
    override_password = request.form.get('override_password', '')

    if password != CONFIRM_CANCEL_PASSWORD:
        flash('비밀번호가 일치하지 않습니다.', 'error')
        return redirect(request.referrer or url_for('index'))
    
    record = users_collection.find_one({'_id': ObjectId(id)}, {'author': 1})

    is_owner = record and record['author'] == current_author
    is_admin = override_password == ADMIN_OVERRIDE_PASSWORD

    if not is_owner and not is_admin:
        flash('타인의 데이터 확정을 취소할 권한이 없습니다.', 'error')
        return redirect(request.referrer or url_for('index'))

    users_collection.update_one({'_id': ObjectId(id)}, {'$set': {'confirmed': False}})
    flash(f'데이터의 확정이 취소되었습니다.', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/download_excel')
def download_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    keyword = request.args.get('keyword')
    author_filter = request.args.get('author_filter')

    # index와 동일한 쿼리 로직 사용
    query = {}
    if start_date and end_date:
        query['work_date'] = {'$gte': start_date, '$lte': end_date}
    if author_filter:
        query['author'] = author_filter
    if keyword:
        query['$or'] = [
            {'client': {'$regex': keyword, '$options': 'i'}},
            {'author': {'$regex': keyword, '$options': 'i'}},
            {'product_name': {'$regex': keyword, '$options': 'i'}},
            {'content': {'$regex': keyword, '$options': 'i'}},
            {'tracking_number': {'$regex': keyword, '$options': 'i'}},
        ]

    cursor = users_collection.find(query).sort([('work_date', 1), ('_id', 1)])
    data_for_excel = list(cursor)

    if data_for_excel:
        df = pd.DataFrame(data_for_excel)
        df['id'] = df['_id'].astype(str) # ObjectId를 문자열로
        df = df.drop(columns=['_id', 'created_at'], errors='ignore')
        if 'attachment' in df.columns:
            df = df.drop(columns=['attachment'])
        # [추가] 데이터프레임의 헤더를 한글로 변경
        df.rename(columns=HEADER_MAP, inplace=True)

        # [추가] 확정여부 값을 '확정'/'미확정' 텍스트로 변경
        df['확정여부'] = df['확정여부'].apply(lambda x: '확정' if x else '미확정')
    else:
        df = pd.DataFrame()

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='정산데이터')

    workbook = writer.book
    worksheet = writer.sheets['정산데이터']

    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # 연한 블루
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for idx, col in enumerate(df.columns, 1):
        column_letter = worksheet.cell(row=1, column=idx).column_letter
        if col in ['작업수량', '박스수량', '금액(단가)', '합계']:
            for cell in worksheet[column_letter][1:]:
                cell.number_format = '#,##0'
        
        # 열 너비 자동 조절
        max_length = 0
        for cell in worksheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    writer.close()
    output.seek(0)
    
    current_month_str = datetime.today().strftime('%Y-%m')
    filename = f"{current_month_str} 정산노트.xlsx"
    encoded_filename = quote(filename)

    return Response(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"})


@app.route('/uploads/<filename>')
def uploaded_file(filename): # 첨부파일 다운로드
    return send_from_directory(ATTACHMENT_DIR, filename)

# 애플리케이션 시작 시 DB 및 디렉토리 초기화
init_dirs()

if __name__ == '__main__':
    app.run(debug=True, port=8000)